from flask import Flask, request, send_file, jsonify
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
from datetime import datetime

app = Flask(__name__)

# ── Colours ───────────────────────────────────────────────────────────────────
DB='1F3864'; MB='2E75B6'; LB='BDD7EE'; LG='F2F2F2'; WH='FFFFFF'
GBG='E2EFDA'; GTX='375623'; ABG='FFF2CC'; ATX='7F6000'; RBG='FCE4D6'; RTX='C00000'
PHDR={'Enterprise':'1F5C99','Commercial In-House':'1E5631','SMB Law':'7F5800','Enterprise AM':'4A235A','SMB AM':'7B1519'}
PROW={'Enterprise':'EBF3FB','Commercial In-House':'EBF7EE','SMB Law':'FDFBE6','Enterprise AM':'F5EEF8','SMB AM':'FDECEA'}
PORD=['Enterprise','Commercial In-House','SMB Law','Enterprise AM','SMB AM']

# Seasonality index — Q-end months surge, Jan/Apr/Jul/Oct are slowest
SEASON = {1:0.72, 2:0.88, 3:2.20, 4:0.72, 5:0.88, 6:2.20, 7:0.72, 8:0.88, 9:2.20, 10:0.72, 11:0.88, 12:2.20}
MONTH_NAMES = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}

# Growth assumptions
NB_GROWTH_RATE   = 0.22   # 22% YoY NB growth (strong pace, recovering from Q1 slow start)
EXP_GROWTH_RATE  = 0.35   # 35% YoY expansion growth (expansion outpacing NB, healthy customer base)
RAMP_COHORT_LIFT = 0.08   # 8% incremental H2 lift from new M1 reps reaching productivity in Q3/Q4
BULL_MULT        = 1.15
BEAR_MULT        = 0.88

def sd(st='thin',co='BFBFBF'): return Side(style=st,color=co)
def bdr(): return Border(left=sd(),right=sd(),top=sd(),bottom=sd())
def fill(hex): return PatternFill('solid',fgColor=hex)
def fnt(bold=False,italic=False,fc='000000',sz=10): return Font(name='Arial',bold=bold,italic=italic,size=sz,color=fc)
def aln(h='center',wrap=False): return Alignment(horizontal=h,vertical='center',wrap_text=wrap)

def C(ws,r,c,v=None,fmt=None,bold=False,italic=False,fc='000000',bg=None,ha='center',wrap=False,sz=10):
    cell=ws.cell(row=r,column=c)
    if v is not None: cell.value=v
    cell.font=fnt(bold,italic,fc,sz)
    if bg: cell.fill=fill(bg)
    cell.alignment=aln(ha,wrap)
    if fmt: cell.number_format=fmt
    cell.border=bdr()
    return cell

def MH(ws,r,c1,c2,txt,bg=None,fc=WH,sz=14,h=34):
    ws.merge_cells(start_row=r,start_column=c1,end_row=r,end_column=c2)
    cell=ws.cell(row=r,column=c1,value=txt)
    cell.font=fnt(True,False,fc,sz); cell.fill=fill(bg or DB); cell.alignment=aln('center')
    ws.row_dimensions[r].height=h

def SEC(ws,r,c1,c2,txt,bg=None,h=20):
    ws.merge_cells(start_row=r,start_column=c1,end_row=r,end_column=c2)
    cell=ws.cell(row=r,column=c1,value=txt)
    cell.font=fnt(True,False,WH,11); cell.fill=fill(bg or MB); cell.alignment=aln('left')
    ws.row_dimensions[r].height=h

def HDR(ws,r,c,txt,bg=None):
    C(ws,r,c,txt,bold=True,bg=bg or DB,fc=WH,ha='center',wrap=True,sz=9)
    ws.row_dimensions[r].height=28

def ATT(ws,r,c,pct):
    bg=GBG if pct>=1.0 else(ABG if pct>=0.75 else RBG)
    fc=GTX if pct>=1.0 else(ATX if pct>=0.75 else RTX)
    C(ws,r,c,pct,'0.0%',bold=True,bg=bg,fc=fc)

def FCAST(ws,r,c,v,is_actual=False):
    """Forecast cell — actual=white/normal, forecast=light yellow italic"""
    bg='FFFDE7' if not is_actual else WH
    fc='000000' if not is_actual else '000000'
    C(ws,r,c,v,'$#,##0',italic=not is_actual,bg=bg,fc=fc)

def GAP(ws,r,h=7): ws.row_dimensions[r].height=h

# ── Forecast engine ───────────────────────────────────────────────────────────
def build_company_monthly_forecast(summary, pod_stats, current_month=3):
    """
    Build month-by-month forecast for full year.
    Jan/Feb: real actuals from podStats.
    Current month: actuals to date only (no projection from 3 days - too noisy).
    Future months: seasonality + growth assumptions applied to 2025 base.
    """
    fy25_nb  = summary['pace2025NB']  * 6
    fy25_exp = summary['pace2025Exp'] * 6
    sv = sum(SEASON.values())

    fy25_monthly_nb  = {m: fy25_nb  * SEASON[m] / sv for m in range(1,13)}
    fy25_monthly_exp = {m: fy25_exp * SEASON[m] / sv for m in range(1,13)}

    # Real actuals direct from podStats - no estimation
    nb_pods  = ['Enterprise', 'Commercial In-House', 'SMB Law']
    exp_pods = [('Enterprise','Enterprise AM'), ('SMB Law','SMB AM')]

    actual_nb = {
        1: round(sum(pod_stats.get(p,{}).get('janNB',0) for p in nb_pods)),
        2: round(sum(pod_stats.get(p,{}).get('febNB',0) for p in nb_pods)),
        3: round(sum(pod_stats.get(p,{}).get('marNB',0) for p in nb_pods)),
    }
    actual_exp = {
        1: round(sum(pod_stats.get(p,{}).get('janExp',0) + pod_stats.get(a,{}).get('janExp',0) for p,a in exp_pods)),
        2: round(sum(pod_stats.get(p,{}).get('febExp',0) + pod_stats.get(a,{}).get('febExp',0) for p,a in exp_pods)),
        3: round(sum(pod_stats.get(p,{}).get('marExp',0) + pod_stats.get(a,{}).get('marExp',0) for p,a in exp_pods)),
    }

    months = {}
    for m in range(1,13):
        is_actual  = m < current_month
        is_partial = m == current_month

        if is_actual:
            # Completed month - use real actuals
            nb  = actual_nb[m]
            exp = actual_exp[m]
        elif is_partial:
            # Current month - show actuals to date, no projection (too few days)
            nb  = actual_nb[m]
            exp = actual_exp[m]
        else:
            # Future month - forecast from 2025 base with growth + ramp cohort lift
            ramp_lift = RAMP_COHORT_LIFT if m >= 7 else 0
            nb  = round(fy25_monthly_nb[m]  * (1 + NB_GROWTH_RATE  + ramp_lift))
            exp = round(fy25_monthly_exp[m] * (1 + EXP_GROWTH_RATE + ramp_lift))

        months[m] = {
            'nb':       nb,
            'exp':      exp,
            'total':    nb + exp,
            'nb_bull':  round(nb  * BULL_MULT) if not is_actual else nb,
            'exp_bull': round(exp * BULL_MULT) if not is_actual else exp,
            'nb_bear':  round(nb  * BEAR_MULT) if not is_actual else nb,
            'exp_bear': round(exp * BEAR_MULT) if not is_actual else exp,
            'is_actual':  is_actual,
            'is_partial': is_partial,
        }
    return months

def build_rep_forecast(rep, current_month=3):
    """Project each rep's full year based on run rate + ramp stage."""
    months_with_data = 2  # Jan + Feb complete
    if rep['ytdRevenue'] == 0:
        monthly_run_rate = 0
    else:
        monthly_run_rate = rep['ytdRevenue'] / months_with_data

    sv = sum(SEASON.values())
    avg_season = sv / 12

    projected = {}
    for m in range(1,13):
        is_actual = m <= 2
        is_partial = m == current_month
        if is_actual:
            rev = rep['janRevenue'] if m==1 else rep['febRevenue']
        elif is_partial:
            rev = rep['marRevenue']
        else:
            season_mult = SEASON[m] / avg_season
            ramp_adj = 1.0
            if rep['rampStatus'] not in ['Fully Ramped']:
                ramp_stage = int(rep['rampStatus'].replace('M','')) if rep['rampStatus'].startswith('M') else 6
                months_ahead = m - current_month
                future_stage = min(ramp_stage + months_ahead, 6)
                ramp_adj = future_stage / 6.0
            rev = round(monthly_run_rate * season_mult * ramp_adj)
        projected[m] = {'rev': rev, 'is_actual': is_actual, 'is_partial': is_partial}

    fy_base = sum(projected[m]['rev'] for m in range(1,13))
    fy_quota = rep.get('ytdQuota',0) * 6  # annualised from 2-month YTD quota
    return projected, fy_base, fy_quota

def build_excel(data):
    summary      = data['summary']
    podStats     = data['podStats']
    repSummaries = data['repSummaries']
    top10Deals   = data.get('top10Deals', data.get('top5Deals',[]))
    thisWeekDeals= data.get('thisWeekDeals',[])
    RL = f"Week Ending {datetime.now().strftime('%b %-d, %Y')}"
    wb = openpyxl.Workbook()
    current_month = datetime.now().month

    # ── SHEET 1: EXECUTIVE SUMMARY ────────────────────────────────────────────
    ws=wb.active; ws.title='Executive Summary'
    ws.sheet_view.showGridLines=False; ws.sheet_view.zoomScale=90
    for col,w in zip('ABCD',[32,16,16,14]): ws.column_dimensions[col].width=w
    MH(ws,1,1,4,'Spellbook Legal — Sales Performance Report',DB,WH,15,38)
    MH(ws,2,1,4,RL,MB,WH,11,22); GAP(ws,3)
    SEC(ws,4,1,4,'  COMPANY OVERVIEW — YTD 2026')
    for i,h in enumerate(['Metric','YTD Actual','YTD Target','Attainment']): HDR(ws,5,i+1,h)
    tot_tgt=summary['totalNBTarget']+summary['totalExpTarget']
    for i,(lbl,act,tgt,att) in enumerate([
        ('Total Revenue',summary['totalRevenue'],tot_tgt,summary['totalRevenue']/tot_tgt if tot_tgt else 0),
        ('New Business',summary['totalNB'],summary['totalNBTarget'],summary['totalNB']/summary['totalNBTarget'] if summary['totalNBTarget'] else 0),
        ('Expansion Revenue',summary['totalExp'],summary['totalExpTarget'],summary['totalExp']/summary['totalExpTarget'] if summary['totalExpTarget'] else 0),
        ('Total Deals',summary['totalDeals'],None,None)]):
        r=6+i; bg=LG if i%2==0 else WH
        C(ws,r,1,lbl,bold=True,bg=bg,ha='left'); C(ws,r,2,act,'$#,##0' if tgt else '#,##0',bg=bg)
        C(ws,r,3,tgt if tgt else '—','$#,##0' if tgt else None,bg=bg)
        if att is not None: ATT(ws,r,4,att)
        else: C(ws,r,4,'—',bg=bg)
        ws.row_dimensions[r].height=18
    GAP(ws,10)
    SEC(ws,11,1,4,'  YEAR-OVER-YEAR — Same Period Est.')
    for i,h in enumerate(['Metric','2026 YTD','2025 Est.','YoY Change']): HDR(ws,12,i+1,h)
    for i,(lbl,v26,v25) in enumerate([
        ('Total Revenue',summary['totalRevenue'],summary['pace2025Total']),
        ('New Business',summary['totalNB'],summary['pace2025NB']),
        ('Expansion',summary['totalExp'],summary['pace2025Exp'])]):
        r=13+i; bg=LG if i%2==0 else WH; chg=(v26-v25)/v25 if v25 else 0
        C(ws,r,1,lbl,bold=True,bg=bg,ha='left'); C(ws,r,2,v26,'$#,##0',bg=bg); C(ws,r,3,v25,'$#,##0',bg=bg)
        C(ws,r,4,chg,'+0.0%;-0.0%;-',bold=True,bg=GBG if chg>=0 else RBG,fc=GTX if chg>=0 else RTX)
        ws.row_dimensions[r].height=18

    # Quick forecast snapshot on exec summary
    GAP(ws,16)
    SEC(ws,17,1,4,'  2026 FULL YEAR FORECAST SNAPSHOT  (see Forecast sheets for detail)')
    for i,h in enumerate(['Scenario','Total Revenue','New Business','Expansion']): HDR(ws,18,i+1,h)
    monthly = build_company_monthly_forecast(summary, podStats, current_month)
    fy_base_nb  = sum(monthly[m]['nb']  for m in range(1,13))
    fy_base_exp = sum(monthly[m]['exp'] for m in range(1,13))
    fy_bull_nb  = sum(monthly[m]['nb_bull']  for m in range(1,13))
    fy_bull_exp = sum(monthly[m]['exp_bull'] for m in range(1,13))
    fy_bear_nb  = sum(monthly[m]['nb_bear']  for m in range(1,13))
    fy_bear_exp = sum(monthly[m]['exp_bear'] for m in range(1,13))
    for i,(lbl,nb,exp,bg) in enumerate([
        ('Base Case',  fy_base_nb, fy_base_exp, LG),
        ('Bull Case (+15%)', fy_bull_nb, fy_bull_exp, GBG),
        ('Bear Case (-12%)', fy_bear_nb, fy_bear_exp, RBG)]):
        r=19+i
        C(ws,r,1,lbl,bold=True,bg=bg,ha='left')
        C(ws,r,2,nb+exp,'$#,##0',bold=True,bg=bg)
        C(ws,r,3,nb,'$#,##0',bg=bg)
        C(ws,r,4,exp,'$#,##0',bg=bg)
        ws.row_dimensions[r].height=18

    # ── SHEET 2: POD PERFORMANCE ──────────────────────────────────────────────
    ws2=wb.create_sheet('Pod Performance')
    ws2.sheet_view.showGridLines=False; ws2.sheet_view.zoomScale=90
    for col,w in zip('ABCDEFGHIJK',[24,14,14,12,14,14,12,14,14,12,10]): ws2.column_dimensions[col].width=w
    MH(ws2,1,1,11,f'Pod Performance — {RL}',DB,WH,14,34); GAP(ws2,2)
    SEC(ws2,3,1,11,'  NEW BUSINESS BY POD — YTD 2026')
    for i,h in enumerate(['Pod','Jan NB','Jan Target','Jan Att%','Feb NB','Feb Target','Feb Att%','Mar NB (partial)','Mar Target','Mar Att%','YTD NB']): HDR(ws2,4,i+1,h)
    for i,pod in enumerate(['Enterprise','Commercial In-House','SMB Law']):
        r=5+i; bg=LG if i%2==0 else WH; ps=podStats[pod]
        jt=ps.get('janNBTarget',0); ft=ps.get('febNBTarget',0); mt=ps.get('marNBTarget',0)
        C(ws2,r,1,pod,bold=True,bg=bg,ha='left')
        C(ws2,r,2,round(ps.get('janNB',0)),'$#,##0',bg=bg); C(ws2,r,3,jt,'$#,##0',bg=bg)
        ATT(ws2,r,4,ps.get('janNB',0)/jt if jt else 0)
        C(ws2,r,5,round(ps.get('febNB',0)),'$#,##0',bg=bg); C(ws2,r,6,ft,'$#,##0',bg=bg)
        ATT(ws2,r,7,ps.get('febNB',0)/ft if ft else 0)
        C(ws2,r,8,round(ps.get('marNB',0)),'$#,##0',bg=bg); C(ws2,r,9,mt,'$#,##0',bg=bg)
        ATT(ws2,r,10,ps.get('marNB',0)/mt if mt else 0)
        C(ws2,r,11,round(ps['newBiz']),'$#,##0',bold=True,bg=bg)
        ws2.row_dimensions[r].height=18
    C(ws2,8,1,'TOTAL NB',bold=True,bg=LB,ha='left')
    C(ws2,8,2,round(sum(podStats[p].get('janNB',0) for p in ['Enterprise','Commercial In-House','SMB Law'])),'$#,##0',bold=True,bg=LB)
    C(ws2,8,5,round(sum(podStats[p].get('febNB',0) for p in ['Enterprise','Commercial In-House','SMB Law'])),'$#,##0',bold=True,bg=LB)
    C(ws2,8,8,round(sum(podStats[p].get('marNB',0) for p in ['Enterprise','Commercial In-House','SMB Law'])),'$#,##0',bold=True,bg=LB)
    C(ws2,8,11,summary['totalNB'],'$#,##0',bold=True,bg=LB)
    for c in [3,4,6,7,9,10]: C(ws2,8,c,'',bg=LB)
    ws2.row_dimensions[8].height=20
    GAP(ws2,9)
    SEC(ws2,10,1,11,'  EXPANSION BY POD — YTD 2026')
    for i,h in enumerate(['Pod','Jan Exp','Jan Target','Jan Att%','Feb Exp','Feb Target','Feb Att%','Mar Exp (partial)','Mar Target','Mar Att%','YTD Exp']): HDR(ws2,11,i+1,h)
    for i,(pod,amPod) in enumerate([('Enterprise','Enterprise AM'),('SMB Law','SMB AM')]):
        r=12+i; bg=LG if i%2==0 else WH; ps=podStats[pod]; am=podStats[amPod]
        totalJanExp=ps.get('janExp',0)+am.get('janExp',0)
        totalFebExp=ps.get('febExp',0)+am.get('febExp',0)
        totalMarExp=ps.get('marExp',0)+am.get('marExp',0)
        totalExp=ps['expansion']+am['expansion']
        jt=ps.get('janExpTarget',0); ft=ps.get('febExpTarget',0); mt=ps.get('marExpTarget',0)
        C(ws2,r,1,pod,bold=True,bg=bg,ha='left')
        C(ws2,r,2,round(totalJanExp),'$#,##0',bg=bg); C(ws2,r,3,jt,'$#,##0',bg=bg)
        ATT(ws2,r,4,totalJanExp/jt if jt else 0)
        C(ws2,r,5,round(totalFebExp),'$#,##0',bg=bg); C(ws2,r,6,ft,'$#,##0',bg=bg)
        ATT(ws2,r,7,totalFebExp/ft if ft else 0)
        C(ws2,r,8,round(totalMarExp),'$#,##0',bg=bg); C(ws2,r,9,mt,'$#,##0',bg=bg)
        ATT(ws2,r,10,totalMarExp/mt if mt else 0)
        C(ws2,r,11,round(totalExp),'$#,##0',bold=True,bg=bg)
        ws2.row_dimensions[r].height=18
    C(ws2,14,1,'TOTAL EXP',bold=True,bg=LB,ha='left')
    C(ws2,14,2,round(sum(podStats[p].get('janExp',0)+podStats[a].get('janExp',0) for p,a in [('Enterprise','Enterprise AM'),('SMB Law','SMB AM')])),'$#,##0',bold=True,bg=LB)
    C(ws2,14,5,round(sum(podStats[p].get('febExp',0)+podStats[a].get('febExp',0) for p,a in [('Enterprise','Enterprise AM'),('SMB Law','SMB AM')])),'$#,##0',bold=True,bg=LB)
    C(ws2,14,8,round(sum(podStats[p].get('marExp',0)+podStats[a].get('marExp',0) for p,a in [('Enterprise','Enterprise AM'),('SMB Law','SMB AM')])),'$#,##0',bold=True,bg=LB)
    C(ws2,14,11,summary['totalExp'],'$#,##0',bold=True,bg=LB)
    for c in [3,4,6,7,9,10]: C(ws2,14,c,'',bg=LB)
    ws2.row_dimensions[14].height=20

    # ── SHEET 3: INDIVIDUAL PERFORMANCE ──────────────────────────────────────
    ws3=wb.create_sheet('Individual Performance')
    ws3.sheet_view.showGridLines=False; ws3.sheet_view.zoomScale=75
    cols3='ABCDEFGHIJKLMNOPQRSTU'
    widths3=[20,26,16,12,12,10,8,10,10,10,9,10,10,9,10,10,9,10,10,12,12]
    for col,w in zip(cols3,widths3): ws3.column_dimensions[col].width=w
    MH(ws3,1,1,21,f'Individual Rep Performance — {RL}',DB,WH,14,34); GAP(ws3,2)
    for i,h in enumerate(['Rep','Role','Pod','YTD Quota','YTD Revenue','Att%','Deals','Avg Deal',
                           'Jan Quota','Jan Rev','Jan Att%','Feb Quota','Feb Rev','Feb Att%',
                           'Mar Quota','Mar Rev','Mar Att%','New Biz','Expansion','2025 FY','YoY']): HDR(ws3,3,i+1,h)
    ws3.freeze_panes='A4'
    row=4
    for pod in PORD:
        reps=sorted([r for r in repSummaries if r['pod']==pod],key=lambda x:x['ytdRevenue'],reverse=True)
        if not reps: continue
        ws3.merge_cells(start_row=row,start_column=1,end_row=row,end_column=21)
        hc=ws3.cell(row=row,column=1,value=f'  {pod.upper()}')
        hc.font=fnt(True,False,WH,10); hc.fill=fill(PHDR.get(pod,MB)); hc.alignment=aln('left')
        ws3.row_dimensions[row].height=18; row+=1
        bg=PROW.get(pod,WH); pR=pQ=pN=pE=pD=0
        for rep in reps:
            att=rep['ytdRevenue']/rep['ytdQuota'] if rep['ytdQuota'] else None
            janA=rep['janRevenue']/rep['janQuota'] if rep.get('janQuota') else None
            febA=rep['febRevenue']/rep['febQuota'] if rep.get('febQuota') else None
            marA=rep['marRevenue']/rep['marQuota'] if rep.get('marQuota') else None
            est25=rep['fy2025Revenue']*2/12 if rep.get('fy2025Revenue') else None
            yoy=(rep['ytdRevenue']-est25)/est25 if est25 else None
            C(ws3,row,1,rep['rep'],bold=True,bg=bg,ha='left',sz=9)
            C(ws3,row,2,rep['role'],bg=bg,ha='left',sz=8)
            C(ws3,row,3,pod,bg=bg,ha='left',sz=9)
            C(ws3,row,4,rep['ytdQuota'],'$#,##0',bg=bg)
            C(ws3,row,5,rep['ytdRevenue'],'$#,##0',bg=bg)
            if att is not None: ATT(ws3,row,6,att)
            else: C(ws3,row,6,'—',bg=bg)
            C(ws3,row,7,rep['ytdDeals'],'#,##0',bg=bg)
            C(ws3,row,8,rep['avgDeal'],'$#,##0',bg=bg)
            C(ws3,row,9,rep.get('janQuota',0),'$#,##0',bg=bg)
            C(ws3,row,10,rep.get('janRevenue',0),'$#,##0',bg=bg)
            if janA is not None: ATT(ws3,row,11,janA)
            else: C(ws3,row,11,'—',bg=bg)
            C(ws3,row,12,rep.get('febQuota',0),'$#,##0',bg=bg)
            C(ws3,row,13,rep.get('febRevenue',0),'$#,##0',bg=bg)
            if febA is not None: ATT(ws3,row,14,febA)
            else: C(ws3,row,14,'—',bg=bg)
            C(ws3,row,15,rep.get('marQuota',0),'$#,##0',bg=bg)
            C(ws3,row,16,rep.get('marRevenue',0),'$#,##0',bg=bg)
            if marA is not None: ATT(ws3,row,17,marA)
            else: C(ws3,row,17,'—',bg=bg)
            C(ws3,row,18,rep['ytdNewBiz'],'$#,##0',bg=bg)
            C(ws3,row,19,rep['ytdExpansion'],'$#,##0',bg=bg)
            C(ws3,row,20,rep['fy2025Revenue'] if rep.get('fy2025Revenue') else '—','$#,##0' if rep.get('fy2025Revenue') else None,bg=bg)
            if yoy is not None: C(ws3,row,21,yoy,'+0.0%;-0.0%;-',bold=True,bg=GBG if yoy>=0 else RBG,fc=GTX if yoy>=0 else RTX)
            else: C(ws3,row,21,'N/A',bg=bg)
            pR+=rep['ytdRevenue']; pQ+=rep['ytdQuota']; pN+=rep['ytdNewBiz']; pE+=rep['ytdExpansion']; pD+=rep['ytdDeals']
            ws3.row_dimensions[row].height=17; row+=1
        C(ws3,row,1,f'{pod} SUBTOTAL',bold=True,bg=LB,ha='left')
        for ci in [2,3]: C(ws3,row,ci,'',bg=LB)
        C(ws3,row,4,pQ,'$#,##0',bold=True,bg=LB); C(ws3,row,5,pR,'$#,##0',bold=True,bg=LB)
        ATT(ws3,row,6,pR/pQ if pQ else 0)
        C(ws3,row,7,pD,'#,##0',bold=True,bg=LB)
        for ci in range(8,22): C(ws3,row,ci,'',bg=LB)
        ws3.row_dimensions[row].height=20; row+=2

    # ── SHEET 4: COMPANY FORECAST ─────────────────────────────────────────────
    ws4=wb.create_sheet('Company Forecast')
    ws4.sheet_view.showGridLines=False; ws4.sheet_view.zoomScale=85
    # 16 columns: Month + NB Base/Bull/Bear + Exp Base/Bull/Bear + Total Base/Bull/Bear + NB Target + Exp Target + Total Target + NB Att + Exp Att
    col_widths=[14,13,13,13,13,13,13,13,13,13,12,12,12,11,11]
    for i,w in enumerate(col_widths): ws4.column_dimensions[chr(65+i)].width=w
    MH(ws4,1,1,15,'2026 Full Year Revenue Forecast — Spellbook Legal',DB,WH,14,34)
    MH(ws4,2,1,15,f'Generated {RL}  |  Base: +22% NB / +35% Exp YoY  |  Bull: +15%  |  Bear: -12%',MB,WH,9,18)
    GAP(ws4,3)

    # Assumptions box
    SEC(ws4,4,1,6,'  FORECAST ASSUMPTIONS')
    for i,(k,v) in enumerate([
        ('NB Growth Rate YoY','22% — strong recovery from Q1 slow start, deal pipeline rebuilding'),
        ('Expansion Growth Rate YoY','35% — expansion outpacing NB; healthy installed base expanding fast'),
        ('H2 Ramp Cohort Lift','8% — 8 new M1 hires reach full productivity in Q3/Q4'),
        ('Seasonality','Q-end surge (Mar/Jun/Sep/Dec = 2.2x avg); Jan/Apr/Jul/Oct slowest at 0.72x'),
        ('Bull Case','Base × 1.15 — strong Q-end closes, ramp cohort overperforms'),
        ('Bear Case','Base × 0.88 — macro headwinds, extended sales cycles in Enterprise')]):
        r=5+i; bg=LG if i%2==0 else WH
        C(ws4,r,1,k,bold=True,bg=bg,ha='left',sz=9)
        ws4.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6)
        cell=ws4.cell(row=r,column=2,value=v); cell.font=fnt(False,True,'444444',9)
        cell.fill=fill(bg); cell.alignment=aln('left'); cell.border=bdr()
        ws4.row_dimensions[r].height=16
    GAP(ws4,11)

    # Monthly forecast table
    SEC(ws4,12,1,15,'  MONTHLY REVENUE FORECAST — BASE / BULL / BEAR')
    hdrs=['Month','NB Base','NB Bull','NB Bear','Exp Base','Exp Bull','Exp Bear','Total Base','Total Bull','Total Bear','NB Target','Exp Target','Total Target','NB Att%','Exp Att%']
    for i,h in enumerate(hdrs): HDR(ws4,13,i+1,h)

    monthly=build_company_monthly_forecast(summary, podStats, current_month)
    sv_total = sum(SEASON.values())
    nb_tgt_monthly_dict  = {m: summary["totalNBTarget"]  * SEASON[m] / sv_total for m in range(1,13)}
    exp_tgt_monthly_dict = {m: summary["totalExpTarget"] * SEASON[m] / sv_total for m in range(1,13)}
    nb_tgt_monthly  = summary["totalNBTarget"]  / 12  # kept for compat
    exp_tgt_monthly = summary["totalExpTarget"] / 12  # kept for compat

    fy_totals = {'nb_base':0,'nb_bull':0,'nb_bear':0,'exp_base':0,'exp_bull':0,'exp_bear':0}
    for m in range(1,13):
        r=14+m-1; d=monthly[m]
        bg = 'E8F4FD' if d['is_actual'] else ('FFFDE7' if d['is_partial'] else WH)
        label = MONTH_NAMES[m] + (' ✓' if d['is_actual'] else (' (partial)' if d['is_partial'] else ' →'))
        C(ws4,r,1,label,bold=d['is_actual'],bg=bg,ha='left')
        C(ws4,r,2,d['nb'],         '$#,##0',bold=d['is_actual'],bg=bg)
        C(ws4,r,3,d['nb_bull'],    '$#,##0',italic=not d['is_actual'],bg='F1FBF1' if not d['is_actual'] else bg)
        C(ws4,r,4,d['nb_bear'],    '$#,##0',italic=not d['is_actual'],bg='FEF5F5' if not d['is_actual'] else bg)
        C(ws4,r,5,d['exp'],        '$#,##0',bold=d['is_actual'],bg=bg)
        C(ws4,r,6,d['exp_bull'],   '$#,##0',italic=not d['is_actual'],bg='F1FBF1' if not d['is_actual'] else bg)
        C(ws4,r,7,d['exp_bear'],   '$#,##0',italic=not d['is_actual'],bg='FEF5F5' if not d['is_actual'] else bg)
        C(ws4,r,8,d['total'],      '$#,##0',bold=d['is_actual'],bg=bg)
        C(ws4,r,9,d['nb_bull']+d['exp_bull'],'$#,##0',italic=not d['is_actual'],bg='F1FBF1' if not d['is_actual'] else bg)
        C(ws4,r,10,d['nb_bear']+d['exp_bear'],'$#,##0',italic=not d['is_actual'],bg='FEF5F5' if not d['is_actual'] else bg)
        C(ws4,r,11,round(nb_tgt_monthly_dict[m]), '$#,##0',bg=LG)
        C(ws4,r,12,round(exp_tgt_monthly_dict[m]),'$#,##0',bg=LG)
        C(ws4,r,13,round(nb_tgt_monthly_dict[m]+exp_tgt_monthly_dict[m]),'$#,##0',bg=LG)
        nb_att  = d["nb"]  / nb_tgt_monthly_dict[m]  if nb_tgt_monthly_dict[m]  else 0
        exp_att = d["exp"] / exp_tgt_monthly_dict[m] if exp_tgt_monthly_dict[m] else 0
        ATT(ws4,r,14,nb_att); ATT(ws4,r,15,exp_att)
        ws4.row_dimensions[r].height=17
        for k in ['nb_base','nb_bull','nb_bear','exp_base','exp_bull','exp_bear']:
            fy_totals[k] += d[k.split('_')[0]] if k.endswith('base') else d[k]

    # Full year totals row
    r=26
    C(ws4,r,1,'FULL YEAR TOTAL',bold=True,bg=LB,ha='left')
    C(ws4,r,2,fy_base_nb,     '$#,##0',bold=True,bg=LB)
    C(ws4,r,3,fy_bull_nb,     '$#,##0',bold=True,bg=GBG)
    C(ws4,r,4,fy_bear_nb,     '$#,##0',bold=True,bg=RBG)
    C(ws4,r,5,fy_base_exp,    '$#,##0',bold=True,bg=LB)
    C(ws4,r,6,fy_bull_exp,    '$#,##0',bold=True,bg=GBG)
    C(ws4,r,7,fy_bear_exp,    '$#,##0',bold=True,bg=RBG)
    C(ws4,r,8,fy_base_nb+fy_base_exp,'$#,##0',bold=True,bg=LB)
    C(ws4,r,9,fy_bull_nb+fy_bull_exp,'$#,##0',bold=True,bg=GBG)
    C(ws4,r,10,fy_bear_nb+fy_bear_exp,'$#,##0',bold=True,bg=RBG)
    C(ws4,r,11,summary['totalNBTarget'], '$#,##0',bold=True,bg=LB)
    C(ws4,r,12,summary['totalExpTarget'],'$#,##0',bold=True,bg=LB)
    C(ws4,r,13,summary['totalNBTarget']+summary['totalExpTarget'],'$#,##0',bold=True,bg=LB)
    ATT(ws4,r,14,fy_base_nb/summary['totalNBTarget'] if summary['totalNBTarget'] else 0)
    ATT(ws4,r,15,fy_base_exp/summary['totalExpTarget'] if summary['totalExpTarget'] else 0)
    ws4.row_dimensions[r].height=22

    # ── SHEET 5: POD FORECAST ─────────────────────────────────────────────────
    ws5=wb.create_sheet('Pod Forecast')
    ws5.sheet_view.showGridLines=False; ws5.sheet_view.zoomScale=85
    for i,w in enumerate([22,12,12,12,12,12,12,12,12,12,12,12,12,14,14]):
        ws5.column_dimensions[chr(65+i)].width=w
    MH(ws5,1,1,15,'2026 Pod-Level Revenue Forecast',DB,WH,14,34)
    MH(ws5,2,1,15,f'Generated {RL}  |  White = Actual  |  Yellow = Forecast',MB,WH,9,18)
    GAP(ws5,3)

    nb_pods  = ['Enterprise','Commercial In-House','SMB Law']
    exp_pods = [('Enterprise','Enterprise AM'),('SMB Law','SMB AM')]

    for section_idx, (section_label, pods_list, is_exp) in enumerate([
        ('NEW BUSINESS BY POD', nb_pods, False),
        ('EXPANSION BY POD',    exp_pods, True)]):
        base_row = 4 + section_idx * 20
        SEC(ws5,base_row,1,15,f'  {section_label}')
        month_hdrs = [MONTH_NAMES[m] for m in range(1,13)]
        for i,h in enumerate(['Pod'] + month_hdrs + ['FY Base','FY Target']): HDR(ws5,base_row+1,i+1,h)

        sv = sum(SEASON.values())
        for pi, pod_item in enumerate(pods_list):
            r = base_row+2+pi; bg = LG if pi%2==0 else WH
            if is_exp:
                pod, amPod = pod_item
                ps = podStats[pod]; am = podStats[amPod]
                actuals_m = {1: ps.get('janExp',0)+am.get('janExp',0),
                             2: ps.get('febExp',0)+am.get('febExp',0),
                             3: ps.get('marExp',0)+am.get('marExp',0)}
                ytd_rev = ps['expansion']+am['expansion']
                fy_tgt  = ps.get('expTarget',0) or am.get('expTarget',0)
                pod_label = pod
                growth = EXP_GROWTH_RATE
                fy25_rev = summary['pace2025Exp'] * 6 * (0.5 if pod=='Enterprise' else 0.5)
            else:
                pod_label = pod_item
                ps = podStats[pod_item]
                actuals_m = {1: ps.get('janNB',0), 2: ps.get('febNB',0), 3: ps.get('marNB',0)}
                ytd_rev = ps['newBiz']
                fy_tgt  = ps.get('nbTarget',0)
                growth = NB_GROWTH_RATE
                fy25_rev = summary['pace2025NB'] * 6 * (0.40 if pod_item=='Enterprise' else (0.35 if pod_item=='Commercial In-House' else 0.25))

            C(ws5,r,1,pod_label,bold=True,bg=bg,ha='left')
            fy_proj = 0
            for m in range(1,13):
                is_actual = m <= 2
                is_partial = m == current_month
                if is_actual:
                    # Completed month - real actuals
                    val = round(actuals_m.get(m, 0))
                    cell_bg = 'E8F4FD'
                elif is_partial:
                    # Current month - actuals to date only, no noisy projection
                    val = round(actuals_m.get(m, 0))
                    cell_bg = 'FFF9C4'
                else:
                    # Future month - forecast
                    ramp_lift = RAMP_COHORT_LIFT if m >= 7 else 0
                    val = round(fy25_rev * SEASON[m] / sv * (1 + growth + ramp_lift))
                    cell_bg = 'FFFDE7'
                C(ws5,r,m+1,val,'$#,##0',italic=not is_actual,bg=cell_bg)
                fy_proj += val
            C(ws5,r,14,fy_proj,'$#,##0',bold=True,bg=bg)
            C(ws5,r,15,fy_tgt,'$#,##0',bold=True,bg=LG)
            ws5.row_dimensions[r].height=17

    # ── SHEET 6: REP FORECAST ─────────────────────────────────────────────────
    ws6=wb.create_sheet('Rep Forecast')
    ws6.sheet_view.showGridLines=False; ws6.sheet_view.zoomScale=75
    for i,w in enumerate([20,14]+[11]*12+[13,13,13,11]):
        ws6.column_dimensions[chr(65+i)].width=w
    MH(ws6,1,1,19,'2026 Rep-Level Revenue Forecast',DB,WH,14,34)
    MH(ws6,2,1,19,f'Generated {RL}  |  Blue = Actual  |  Yellow = Forecast  |  Run-rate projected with seasonality + ramp stage',MB,WH,9,18)
    GAP(ws6,3)
    month_hdrs=[MONTH_NAMES[m] for m in range(1,13)]
    for i,h in enumerate(['Rep','Pod']+month_hdrs+['FY Projected','FY Quota','Quota Att%','Risk Flag']): HDR(ws6,4,i+1,h)
    ws6.freeze_panes='A5'

    row=5
    for pod in PORD:
        reps=sorted([r for r in repSummaries if r['pod']==pod],key=lambda x:x['ytdRevenue'],reverse=True)
        if not reps: continue
        ws6.merge_cells(start_row=row,start_column=1,end_row=row,end_column=19)
        hc=ws6.cell(row=row,column=1,value=f'  {pod.upper()}')
        hc.font=fnt(True,False,WH,10); hc.fill=fill(PHDR.get(pod,MB)); hc.alignment=aln('left')
        ws6.row_dimensions[row].height=18; row+=1

        for rep in reps:
            projected, fy_proj, fy_quota = build_rep_forecast(rep, current_month)
            # Annualise quota properly
            ramp_row = data.get('ramp',[])
            role_quota = next((x.get('Monthly Quota 2026',0) for x in ramp_row if x.get('Role')==rep['role']),0)
            fy_quota_full = role_quota * 12
            att = fy_proj / fy_quota_full if fy_quota_full else 0

            # Risk flag
            if rep['rampStatus'] not in ['Fully Ramped'] and rep['ytdRevenue'] == 0:
                flag = '⚠ No revenue yet'
                flag_bg = RBG
            elif att < 0.6:
                flag = '🔴 At risk'
                flag_bg = RBG
            elif att < 0.8:
                flag = '🟡 Watch'
                flag_bg = ABG
            elif att >= 1.0:
                flag = '🟢 On track'
                flag_bg = GBG
            else:
                flag = '🟢 On track'
                flag_bg = GBG

            bg = PROW.get(pod, WH)
            C(ws6,row,1,rep['rep'],bold=True,bg=bg,ha='left',sz=9)
            C(ws6,row,2,pod,bg=bg,ha='left',sz=9)
            for m in range(1,13):
                d=projected[m]
                cell_bg='E8F4FD' if d['is_actual'] else ('FFFDE7' if d['is_partial'] else bg)
                C(ws6,row,m+2,d['rev'],'$#,##0',italic=not d['is_actual'],bg=cell_bg)
            C(ws6,row,15,fy_proj,'$#,##0',bold=True,bg=bg)
            C(ws6,row,16,fy_quota_full,'$#,##0',bg=LG)
            ATT(ws6,row,17,att)
            C(ws6,row,18,rep['rampStatus'],bg=bg)
            C(ws6,row,19,flag,bold=True,bg=flag_bg,ha='left',sz=9)
            ws6.row_dimensions[row].height=17; row+=1
        row+=1

    # ── SHEET 7: HIRING IMPACT ────────────────────────────────────────────────
    ws7=wb.create_sheet('Hiring Impact')
    ws7.sheet_view.showGridLines=False; ws7.sheet_view.zoomScale=85
    for i,w in enumerate([24,16,12,12,12,12,12,12,12,12,12,12,12,14,14]):
        ws7.column_dimensions[chr(65+i)].width=w
    MH(ws7,1,1,15,'2026 New Hire Ramp & Revenue Impact',DB,WH,14,34)
    MH(ws7,2,1,15,'Shows incremental revenue contribution from M1 hires as they ramp through the year',MB,WH,9,18)
    GAP(ws7,3)
    SEC(ws7,4,1,15,'  NEW M1 HIRES — RAMP CONTRIBUTION BY MONTH')
    for i,h in enumerate(['Rep','Role']+[MONTH_NAMES[m] for m in range(1,13)]+['FY Total','Full Quota']): HDR(ws7,5,i+1,h)

    new_hires = [r for r in repSummaries if r.get('rampStatus','').startswith('M') and r['rampStatus'] != 'Fully Ramped']
    row=6
    cohort_by_month = {m:0 for m in range(1,13)}
    for rep in new_hires:
        projected, fy_proj, _ = build_rep_forecast(rep, current_month)
        ramp_row_data = data.get('ramp',[])
        role_quota = next((x.get('Monthly Quota 2026',0) for x in ramp_row_data if x.get('Role')==rep['role']),0)
        fy_full_quota = role_quota * 12
        bg=LG if row%2==0 else WH
        C(ws7,row,1,rep['rep'],bold=True,bg=bg,ha='left',sz=9)
        C(ws7,row,2,rep['role'],bg=bg,ha='left',sz=8)
        for m in range(1,13):
            d=projected[m]
            cell_bg='E8F4FD' if d['is_actual'] else 'FFFDE7'
            C(ws7,row,m+2,d['rev'],'$#,##0',italic=not d['is_actual'],bg=cell_bg)
            cohort_by_month[m]+=d['rev']
        C(ws7,row,15,fy_proj,'$#,##0',bold=True,bg=bg)
        C(ws7,row,16,fy_full_quota,'$#,##0',bg=LG)
        ws7.row_dimensions[row].height=17; row+=1

    # Cohort total row
    GAP(ws7,row); row+=1
    C(ws7,row,1,'COHORT TOTAL',bold=True,bg=LB,ha='left')
    C(ws7,row,2,'',bg=LB)
    cohort_fy=0
    for m in range(1,13):
        C(ws7,row,m+2,cohort_by_month[m],'$#,##0',bold=True,bg=LB)
        cohort_fy+=cohort_by_month[m]
    C(ws7,row,15,cohort_fy,'$#,##0',bold=True,bg=LB)
    C(ws7,row,16,'',bg=LB)
    ws7.row_dimensions[row].height=20; row+=2

    # Commentary
    SEC(ws7,row,1,8,'  HIRING IMPACT ANALYSIS')
    row+=1
    h2_lift=sum(cohort_by_month[m] for m in range(7,13))
    h1_rev =sum(cohort_by_month[m] for m in range(1,7))
    notes=[
        f'H1 cohort contribution: ${h1_rev:,.0f} — reps still ramping, limited contribution',
        f'H2 cohort contribution: ${h2_lift:,.0f} — reps reach M3-M6, meaningful productivity',
        f'Full year cohort revenue: ${cohort_fy:,.0f} vs what they would contribute at full quota',
        f'These {len(new_hires)} reps represent the growth investment — track weekly to ensure ramp is on schedule',
        'Early indicator: Any M1 rep with $0 revenue by end of M2 needs immediate manager attention'
    ]
    for note in notes:
        ws7.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8)
        cell=ws7.cell(row=row,column=1,value=f'  • {note}')
        cell.font=fnt(False,False,'333333',9); cell.alignment=aln('left')
        cell.fill=fill(LG if row%2==0 else WH); cell.border=bdr()
        ws7.row_dimensions[row].height=16; row+=1

    # ── SHEET 8: TOP 10 DEALS ─────────────────────────────────────────────────
    ws8=wb.create_sheet('Top 10 Deals YTD')
    ws8.sheet_view.showGridLines=False
    for col,w in zip('ABCDE',[50,22,20,16,20]): ws8.column_dimensions[col].width=w
    MH(ws8,1,1,5,f'Top 10 Deals YTD — {RL}',DB,WH,14,34); GAP(ws8,2)
    for i,h in enumerate(['Deal Name','Owner','Pipeline','Amount','Revenue Start Date']): HDR(ws8,3,i+1,h)
    for i,deal in enumerate(top10Deals):
        r=4+i; bg=LG if i%2==0 else WH
        C(ws8,r,1,deal['dealname'],bg=bg,ha='left'); C(ws8,r,2,deal['owner'],bg=bg)
        C(ws8,r,3,deal['pipeline'],bg=bg); C(ws8,r,4,deal['amount'],'$#,##0',bold=True,bg=bg)
        C(ws8,r,5,deal['revenue_start_date'],bg=bg); ws8.row_dimensions[r].height=18

    # ── SHEET 9: THIS WEEK'S DEALS ────────────────────────────────────────────
    ws9=wb.create_sheet("This Week's Deals")
    ws9.sheet_view.showGridLines=False
    for col,w in zip('ABCDE',[50,22,20,16,20]): ws9.column_dimensions[col].width=w
    MH(ws9,1,1,5,f"This Week's Deals (Roster Only) — {RL}",DB,WH,14,34); GAP(ws9,2)
    for i,h in enumerate(['Deal Name','Owner','Pipeline','Amount','Revenue Start Date']): HDR(ws9,3,i+1,h)
    if thisWeekDeals:
        for i,deal in enumerate(thisWeekDeals):
            r=4+i; bg=LG if i%2==0 else WH
            C(ws9,r,1,deal['dealname'],bg=bg,ha='left'); C(ws9,r,2,deal['owner'],bg=bg)
            C(ws9,r,3,deal['pipeline'],bg=bg); C(ws9,r,4,deal['amount'],'$#,##0',bold=True,bg=bg)
            C(ws9,r,5,deal['revenue_start_date'],bg=bg); ws9.row_dimensions[r].height=18
    else:
        ws9.merge_cells('A4:E4')
        c=ws9.cell(row=4,column=1,value='No roster deals with revenue start date in the last 7 days.')
        c.font=fnt(False,True,'888888',10); c.alignment=aln('left'); ws9.row_dimensions[4].height=20

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

@app.route('/health',methods=['GET'])
def health():
    return jsonify({'status':'ok','service':'spellbook-excel-report','version':'2.0-forecast'})

@app.route('/generate-report',methods=['POST'])
def generate_report():
    try:
        data=request.get_json(force=True)
        if not data: return jsonify({'error':'No JSON body received'}),400
        buf=build_excel(data)
        filename=f"Sales_Report_{datetime.now().strftime('%b%-d_%Y')}.xlsx"
        return send_file(buf,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,download_name=filename)
    except Exception as e:
        import traceback
        return jsonify({'error':str(e),'trace':traceback.format_exc()}),500

if __name__=='__main__':
    app.run(host='0.0.0.0',port=8080)
